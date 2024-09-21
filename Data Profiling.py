# Databricks notebook source
!pip install ydata-profiling[pyspark]
!pip install --upgrade Pillow
!pip install openpyxl
%restart_python 

# COMMAND ----------

from pyspark.sql import SparkSession
from pyspark.sql import functions as F
from pyspark.sql import Window as W

spark = SparkSession.builder.appName("IPL Dataprofiling").getOrCreate()

# COMMAND ----------

files = ["Team", "Player", "Match", "Player_match", "Ball_By_Ball"]
dfs = {file: spark.read.csv(f"/mnt/ipl-data/raw/{file}.csv", header=True) for file in files}

# COMMAND ----------

from ydata_profiling import ProfileReport
for file, df in dfs.items():
    report = ProfileReport(
        df,
        title=f"Profiling {file} DataFrame",
        infer_dtypes=False,
        interactions=None,
        missing_diagrams=None,
    )
    report.to_file(f"DP/{file}.html") 

# COMMAND ----------

from functools import reduce

def data_profile(df):
    result_df_arr = []
    for col in df.columns:
        agg = (
            df
            .withColumn("attribute",F.lit(col))
            .groupBy("attribute")
            .agg(
            F.count("*").alias("count"),
            F.countDistinct(col).alias("unique"),
            F.concat(F.round(F.countDistinct(col) * 100 / F.count("*"), 2), F.lit("%")).alias("unique %"),
            F.sum(F.when(F.col(col).isNull(), 1).otherwise(0)).alias("missing"),
            F.concat(F.round(F.sum(F.when(F.col(col).isNull(), 1).otherwise(0)) * 100 / F.count("*"), 2), F.lit("%")).alias("missing %"),
            F.min(col).alias("min"),
            F.max(col).alias("max"),
            F.round(F.avg(col).alias("avg"), 2),
            F.sum(col).alias("sum"),
            F.round(F.sum(F.length(col)) / F.count("*"), 2).alias("avg len"),
            )
        )
        frequency = (
            df
            .groupBy(col)
            .count()
            .sort("count", ascending=False)
            .limit(1)
            .withColumnRenamed(col, "top")
            .withColumnRenamed("count", "top_frequency")
        )

        final_df = (
            agg
            .crossJoin(frequency)
            .withColumn("top freq %", F.round((F.col("top_frequency") / F.col("count")) * 100, 2))
        )
        result_df_arr.append(final_df)
    return reduce(lambda x, y: x.union(y), result_df_arr)

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font

# def excel_formatting(workbook_path):
#     workbook = load_workbook(workbook_path)
#     thin_border = Border(
#        left=Side(style='thin'),
#        right=Side(style='thin'),
#        top=Side(style='thin'),
#        bottom=Side(style='thin')
#     )
#     header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
#     header_font = Font(bold=True,color='FFFFFF') 
#     for sheet_name in workbook.sheetnames:
#        worksheet = workbook[sheet_name]
#        for row in worksheet.iter_rows():
#            for cell in row:
#                cell.border = thin_border
#        for cell in worksheet[1]:
#            cell.fill = header_fill
#            cell.font = header_font
#     workbook.save(workbook_path)

def excel_formatting(workbook_path):
    workbook = load_workbook(workbook_path)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(bold=True, color='FFFFFF')

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try: 
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        for row in worksheet.iter_rows():
            max_height = 15
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    max_height = max(max_height, (cell.value.count('\n') + 1) * 15)
            worksheet.row_dimensions[cell.row].height = max_height

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

    workbook.save(workbook_path)


# COMMAND ----------

import pandas as pd

with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
    for file, df in dfs.items():
        data_profile(df).toPandas().to_excel(writer, sheet_name=file, index=False)

excel_formatting("output.xlsx")
